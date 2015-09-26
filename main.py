import logging
logging.basicConfig(filename='info.log', level=logging.DEBUG)
logging.info("first.py started & logging imported")
logging.info("loglevel: DEBUG")

from openpyxl import *
logging.info("openpyxl imported")

import glob
import os
logging.info("glob & os imported")

totaldata = []
os.chdir(r"C:\Users\Markus Laaksonen\PycharmProjects\FirstOne")
for data in glob.glob("Osmosis*.xlsx"):
    totaldata.append(data)
logging.info("Total matched files listed")

listiterations = []
for n in range(0, len(totaldata), 1):
    r = totaldata[n]
    s = ''.join(x for x in r if x.isdigit())
    listiterations.append(s)
logging.info("Iterations of files listed")

newestiteration = int(listiterations[-1])
file = "Osmosis Lab Raw Data_" + str(newestiteration) + ".xlsx"
logging.info("Found newest iteration of workbook")

wb = load_workbook(file)
logging.info('Workbook "' + file + '" loaded')

wb._active_sheet_index = 0
logging.info("Marked first sheet as active")

ws = wb.active
logging.info("Active sheet loaded")

# print(ws['D5'].value)

print("----RESULTS FOR 1ST GROUP----")
logging.info("Asking user for input/results")
for n in range(2, 10, 2):  # Ask for values for every beaker in group 1
    item = ws['A' + str(n)]
    inp = input("Weight of: " + str(item.value) + ": ")
    if inp == "":
        ws['F' + str(n)] = inp
    else:
        ws['F' + str(n)] = float(inp.replace(',', '.'))
logging.info("Got values for group 1")

print("----RESULTS FOR 2ND GROUP----")
for n in range(3, 10, 2):  # Ask for values for every beaker in group 2
    item = ws['A' + str(n)]
    inp = input("Weight of: " + str(item.value) + ": ")
    if inp == "":
        ws['F' + str(n)] = inp
    else:
        ws['F' + str(n)] = float(inp.replace(',', '.'))
logging.info("Got values for group 2")
print("----RESULT SAVED: SUCCESS----")

logging.info("Saving file - START")
filen, filef = file.split('.')
logging.debug("file split into: " + filen + " AND " + filef)
filen1, filen2 = filen.split('_')
logging.debug("filename split further into: " + filen1 + " AND " + filen2)
iteration = int(filen2) + 1
logging.debug("Calculated newest iteration")
wb.save(filen1 + "_" + str(iteration) + "." + filef)
logging.info("Saving file - FINISHED")

finalname = filen1 + "_" + str(iteration) + "." + filef
logging.info("Saved as new file: " + finalname)
logging.info("Old file: " + file)
